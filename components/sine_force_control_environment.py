# -*- coding: utf-8 -*-
"""
Closed-loop force-controlled continuous sine sweep environment.

Drives an electromagnetic shaker with a phase-continuous sine sweep whose
*amplitude* (not frequency) is continuously adjusted so a measured force
channel tracks a target force amplitude, per the architecture:

    SineSweepGenerator (output)          ForceTrackingEstimator (input)
            |                                      |
            v                                      v
    data_out_queue <-- drive_amplitude --  ForceAmplitudeController
                                                     ^
                                          control_update_period_s (sample-
                                          counted, independent of DAQ block
                                          size / samples_per_frame)

This environment follows the same single-process, raw-block data flow as
``components/time_environment.py`` (documented there as the template for new
control types): ``data_in_queue`` delivers raw, already-calibrated time
blocks (no FFT), and ``data_out_queue`` receives raw output blocks -- there
is no FFT-bin dependency anywhere in the control path (see
``force_tracking_estimator.py``).

Because acquisition (input) and generation (output) are pipelined with some
latency, the phase/frequency reference used to demodulate an *incoming*
force block is produced by a second ``SineSweepGenerator`` instance
(``_input_phase_generator``) configured identically to the output generator
but advanced by *input* samples consumed rather than *output* samples
generated. Since the frequency law is a pure function of elapsed time (see
``sine_sweep_generator.py``), both generators track the same trajectory,
offset by only the roughly-constant pipeline latency -- which does not
corrupt the amplitude estimate (see ``force_tracking_estimator.py``
docstring), provided the sweep is slow relative to that latency (the
existing quasi-stationary-sweep assumption of this whole design).

Safety notes (see also the module-level report delivered alongside this
file):

* Per-channel Warning/Abort levels are configured in the main DAQ channel
  table and enforced by the existing, unmodified
  ``components/acquisition.py`` mechanism -- this environment does not
  duplicate that. It only adds a *second*, independent hard safety net on
  the drive voltage itself (``abort_drive_v``), since that is a quantity
  internal to this environment that the generic channel-table mechanism has
  no visibility into.
* ``max_drive_v`` is the *control* limit (the controller will never request
  more). ``abort_drive_v`` is the *safety* limit: if the commanded drive
  amplitude ever exceeds it (which should be impossible if
  ``abort_drive_v > max_drive_v`` and the controller is implemented
  correctly, but is checked independently as defense in depth), the
  environment immediately mutes its output and requests a global hardware
  stop. There is no automatic resumption after an abort.
* NOT implemented (limitation, see report): DAQ input overrange/clipping
  detection and DAQ output fault detection are not implemented anywhere in
  the existing Rattlesnake codebase (verified during analysis) and are not
  added here, as doing so would require changes to the shared
  acquisition/output layer well beyond this environment's scope.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import copy
import json
import os
import time
import traceback
import multiprocessing as mp
from multiprocessing.queues import Queue

import numpy as np
import netCDF4 as nc4
import openpyxl
from qtpy import QtWidgets, uic

from .abstract_environment import AbstractEnvironment, AbstractMetadata, AbstractUI
from .utilities import DataAcquisitionParameters, VerboseMessageQueue, GlobalCommands
from .ui_utilities import multiline_plotter
from .environments import (ControlTypes, environment_definition_ui_paths,
                           environment_run_ui_paths)
from .sine_sweep_generator import SineSweepGenerator
from .force_tracking_estimator import ForceTrackingEstimator
from .force_amplitude_controller import ForceAmplitudeController, ControllerStatus
from .force_target_specification import ConstantForceTarget
from .feedforward_map import FeedforwardMap, compose_drive_amplitude

control_type = ControlTypes.SINE_FORCE
WAIT_TIME = 0.001
MAX_PLOT_SAMPLES = 2000
# Floor applied to the adaptive tracking bandwidth (matches the Tracking
# Bandwidth spinbox's own minimum) -- purely a numerical safety net, not a
# meaningful physical limit, since the configured sweep frequency itself
# never goes below f_start/f_stop (>0, validated at environment setup).
MIN_ADAPTIVE_TRACKING_BANDWIDTH_HZ = 0.01
# Where the definition tab's "last used settings" (see
# SineForceControlUI._save_last_used_settings/_restore_last_used_settings)
# are remembered across application restarts. A dotfile under the user's
# home directory -- not inside the repo/install tree -- so it is never
# picked up by version control and works the same regardless of where the
# application happens to be installed/run from (no existing user-config-
# directory convention exists elsewhere in this codebase to match instead).
LAST_USED_SETTINGS_PATH = os.path.join(
    os.path.expanduser('~'), '.rattlesnake', 'sine_force_control_last_settings.json')

# Feedforward learning knobs that are deliberately *not* exposed as UI
# fields (unlike Learning Rate / Feedforward Min-Max / Bins-per-Decade,
# which materially change test behavior and are worth exposing) -- these
# are robustness internals of components.feedforward_map.FeedforwardMap
# with defaults that should rarely need changing; see that module's
# docstring for what each one protects against.
FEEDFORWARD_OUTLIER_REJECT_RATIO = 4.0
FEEDFORWARD_OUTLIER_REJECT_MIN_OBSERVATIONS = 2.0
FEEDFORWARD_OUTLIER_REJECT_PERSISTENCE = 3.0
FEEDFORWARD_MAX_RELATIVE_STEP_PER_UPDATE = 0.3
FEEDFORWARD_MAX_OBSERVATIONS_CAP = 50.0
# Not wired to the UI: a single shared A_FF(f) curve is learned from both
# sweep directions by default (simpler, more sample-efficient). Flipping
# this to True (and adding UI for it) is the extension point mentioned in
# the design notes for separately learning A_FF_up(f)/A_FF_down(f) should a
# real hysteresis/sweep-rate effect ever be observed in practice.
FEEDFORWARD_SEPARATE_DIRECTION = False
# Per-update step limit (dimensionless) on the trim gain `g` in
# u_total = A_FF(f) * g -- the feedforward-mode analogue of Max Drive
# Change/Update (max_drive_step_v), which instead bounds the *composed*
# command directly (see SineForceControlEnvironment._compose_drive_amplitude).
FEEDFORWARD_TRIM_GAIN_MAX_STEP = 0.5

SWEEP_TYPE_UI_TO_INTERNAL = {'Linear': 'linear', 'Logarithmic': 'logarithmic'}
SWEEP_TYPE_INTERNAL_TO_UI = {v: k for k, v in SWEEP_TYPE_UI_TO_INTERNAL.items()}
DIRECTION_UI_TO_INTERNAL = {'Up': 'up', 'Down': 'down'}
DIRECTION_INTERNAL_TO_UI = {v: k for k, v in DIRECTION_UI_TO_INTERNAL.items()}


class SineForceControlQueues:
    """Set of queues used by the Sine Force Control environment."""

    def __init__(self,
                 environment_command_queue: VerboseMessageQueue,
                 gui_update_queue: mp.queues.Queue,
                 controller_communication_queue: VerboseMessageQueue,
                 data_in_queue: mp.queues.Queue,
                 data_out_queue: mp.queues.Queue,
                 log_file_queue: VerboseMessageQueue):
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.log_file_queue = log_file_queue


class SineForceControlParameters(AbstractMetadata):
    """Storage container for parameters used by the Sine Force Control environment."""

    def __init__(self,
                 sample_rate: float,
                 output_sample_rate: float,
                 sweep_type: str,
                 f_start: float,
                 f_stop: float,
                 sweep_rate: float,
                 direction: str,
                 repeat: bool,
                 num_sweeps: int,
                 alternate_direction: bool,
                 force_channel_index: int,
                 target_force: float,
                 force_floor: float,
                 tracking_bandwidth_hz: float,
                 adaptive_tracking_bandwidth: bool,
                 tracking_cycles: float,
                 controller_alpha: float,
                 control_update_period_s: float,
                 initial_drive_v: float,
                 max_drive_v: float,
                 abort_drive_v: float,
                 max_drive_step_v: float,
                 ramp_time_s: float,
                 pre_dwell_time_s: float,
                 feedforward_enabled: bool = False,
                 feedforward_learning_rate: float = 0.2,
                 feedforward_min_v: float = 0.05,
                 feedforward_max_v: float = 1.25,
                 feedforward_trim_gain_max: float = 3.0,
                 feedforward_bins_per_decade: float = 10.0,
                 feedforward_file: str = '',
                 feedforward_load_on_start: bool = False,
                 feedforward_save_on_finish: bool = False):
        self.sample_rate = sample_rate
        self.output_sample_rate = output_sample_rate
        self.sweep_type = sweep_type
        self.f_start = f_start
        self.f_stop = f_stop
        self.sweep_rate = sweep_rate
        self.direction = direction
        self.repeat = repeat
        self.num_sweeps = num_sweeps
        self.alternate_direction = alternate_direction
        self.force_channel_index = force_channel_index
        self.target_force = target_force
        self.force_floor = force_floor
        self.tracking_bandwidth_hz = tracking_bandwidth_hz
        self.adaptive_tracking_bandwidth = adaptive_tracking_bandwidth
        self.tracking_cycles = tracking_cycles
        self.controller_alpha = controller_alpha
        self.control_update_period_s = control_update_period_s
        self.initial_drive_v = initial_drive_v
        self.max_drive_v = max_drive_v
        self.abort_drive_v = abort_drive_v
        self.max_drive_step_v = max_drive_step_v
        self.ramp_time_s = ramp_time_s
        self.pre_dwell_time_s = pre_dwell_time_s
        self.feedforward_enabled = feedforward_enabled
        self.feedforward_learning_rate = feedforward_learning_rate
        self.feedforward_min_v = feedforward_min_v
        self.feedforward_max_v = feedforward_max_v
        self.feedforward_trim_gain_max = feedforward_trim_gain_max
        self.feedforward_bins_per_decade = feedforward_bins_per_decade
        self.feedforward_file = feedforward_file
        self.feedforward_load_on_start = feedforward_load_on_start
        self.feedforward_save_on_finish = feedforward_save_on_finish

    def store_to_netcdf(self, netcdf_group_handle: nc4._netCDF4.Group):
        """Stores parameters and creates the diagnostic time-series variables."""
        netcdf_group_handle.sweep_type = self.sweep_type
        netcdf_group_handle.f_start = self.f_start
        netcdf_group_handle.f_stop = self.f_stop
        netcdf_group_handle.sweep_rate = self.sweep_rate
        netcdf_group_handle.direction = self.direction
        netcdf_group_handle.repeat = 1 if self.repeat else 0
        netcdf_group_handle.num_sweeps = self.num_sweeps
        netcdf_group_handle.alternate_direction = 1 if self.alternate_direction else 0
        netcdf_group_handle.force_channel_index = self.force_channel_index
        netcdf_group_handle.target_force = self.target_force
        netcdf_group_handle.force_unit = 'peak'
        netcdf_group_handle.force_floor = self.force_floor
        netcdf_group_handle.tracking_bandwidth_hz = self.tracking_bandwidth_hz
        netcdf_group_handle.adaptive_tracking_bandwidth = 1 if self.adaptive_tracking_bandwidth else 0
        netcdf_group_handle.tracking_cycles = self.tracking_cycles
        netcdf_group_handle.controller_alpha = self.controller_alpha
        netcdf_group_handle.control_update_period_s = self.control_update_period_s
        netcdf_group_handle.initial_drive_v = self.initial_drive_v
        netcdf_group_handle.max_drive_v = self.max_drive_v
        netcdf_group_handle.abort_drive_v = self.abort_drive_v
        netcdf_group_handle.max_drive_step_v = self.max_drive_step_v
        netcdf_group_handle.ramp_time_s = self.ramp_time_s
        netcdf_group_handle.pre_dwell_time_s = self.pre_dwell_time_s
        netcdf_group_handle.feedforward_enabled = 1 if self.feedforward_enabled else 0
        netcdf_group_handle.feedforward_learning_rate = self.feedforward_learning_rate
        netcdf_group_handle.feedforward_min_v = self.feedforward_min_v
        netcdf_group_handle.feedforward_max_v = self.feedforward_max_v
        netcdf_group_handle.feedforward_trim_gain_max = self.feedforward_trim_gain_max
        netcdf_group_handle.feedforward_bins_per_decade = self.feedforward_bins_per_decade
        netcdf_group_handle.feedforward_file = self.feedforward_file
        netcdf_group_handle.feedforward_load_on_start = 1 if self.feedforward_load_on_start else 0
        netcdf_group_handle.feedforward_save_on_finish = 1 if self.feedforward_save_on_finish else 0
        netcdf_group_handle.createDimension('control_updates', None)
        netcdf_group_handle.createVariable('time', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('instantaneous_frequency', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('force_target', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('force_amplitude_measured', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('relative_force_error', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('drive_amplitude_command', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('controller_state', str, ('control_updates',))
        netcdf_group_handle.createVariable('controller_saturated', 'i1', ('control_updates',))
        netcdf_group_handle.createVariable('estimator_valid', 'i1', ('control_updates',))
        # Feedforward learning diagnostics (see components/feedforward_map.py).
        # feedforward_value/feedback_correction_pct are NaN throughout when
        # feedforward_enabled=0 -- the fast loop then drives
        # drive_amplitude_command directly, as before this feature existed.
        netcdf_group_handle.createVariable('feedforward_value', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('feedback_correction_pct', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('feedforward_confidence', 'f8', ('control_updates',))
        netcdf_group_handle.createVariable('feedforward_learning_applied', 'i1', ('control_updates',))
        netcdf_group_handle.createVariable('sweep_direction', str, ('control_updates',))
        netcdf_group_handle.createVariable('sweep_number', 'i4', ('control_updates',))

    @classmethod
    def from_ui(cls, ui: 'SineForceControlUI') -> 'SineForceControlParameters':
        widget = ui.definition_widget
        sample_rate = ui.data_acquisition_parameters.sample_rate
        output_sample_rate = sample_rate * ui.data_acquisition_parameters.output_oversample
        return cls(
            sample_rate=sample_rate,
            output_sample_rate=output_sample_rate,
            sweep_type=SWEEP_TYPE_UI_TO_INTERNAL[widget.sweep_type_selector.currentText()],
            f_start=widget.start_frequency_selector.value(),
            f_stop=widget.stop_frequency_selector.value(),
            sweep_rate=widget.sweep_rate_selector.value(),
            direction=DIRECTION_UI_TO_INTERNAL[widget.direction_selector.currentText()],
            repeat=widget.repeat_sweep_checkbox.isChecked(),
            num_sweeps=widget.num_sweeps_selector.value(),
            alternate_direction=widget.alternate_direction_checkbox.isChecked(),
            force_channel_index=widget.force_channel_selector.currentData(),
            target_force=widget.target_force_selector.value(),
            force_floor=widget.force_floor_selector.value(),
            tracking_bandwidth_hz=widget.tracking_bandwidth_selector.value(),
            adaptive_tracking_bandwidth=widget.adaptive_tracking_bandwidth_checkbox.isChecked(),
            tracking_cycles=widget.tracking_cycles_selector.value(),
            controller_alpha=widget.controller_alpha_selector.value(),
            control_update_period_s=widget.control_update_period_selector.value(),
            initial_drive_v=widget.initial_drive_selector.value(),
            max_drive_v=widget.max_drive_selector.value(),
            abort_drive_v=widget.abort_drive_selector.value(),
            max_drive_step_v=widget.max_drive_step_selector.value(),
            ramp_time_s=widget.ramp_time_selector.value(),
            pre_dwell_time_s=widget.pre_dwell_time_selector.value(),
            feedforward_enabled=widget.feedforward_enabled_checkbox.isChecked(),
            feedforward_learning_rate=widget.feedforward_learning_rate_selector.value(),
            feedforward_min_v=widget.feedforward_min_selector.value(),
            feedforward_max_v=widget.feedforward_max_selector.value(),
            feedforward_trim_gain_max=widget.feedforward_trim_gain_max_selector.value(),
            feedforward_bins_per_decade=widget.feedforward_bins_per_decade_selector.value(),
            feedforward_file=widget.feedforward_file_selector.text(),
            feedforward_load_on_start=widget.feedforward_load_on_start_checkbox.isChecked(),
            feedforward_save_on_finish=widget.feedforward_save_on_finish_checkbox.isChecked(),
        )


class SineForceControlUI(AbstractUI):
    """User interface for the Sine Force Control environment."""

    def __init__(self,
                 environment_name: str,
                 definition_tabwidget: QtWidgets.QTabWidget,
                 system_id_tabwidget: QtWidgets.QTabWidget,
                 test_predictions_tabwidget: QtWidgets.QTabWidget,
                 run_tabwidget: QtWidgets.QTabWidget,
                 environment_command_queue: VerboseMessageQueue,
                 controller_communication_queue: VerboseMessageQueue,
                 log_file_queue: Queue):
        super().__init__(environment_name,
                          environment_command_queue, controller_communication_queue,
                          log_file_queue)
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[control_type], self.definition_widget)
        definition_tabwidget.addTab(self.definition_widget, self.environment_name)
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[control_type], self.run_widget)
        run_tabwidget.addTab(self.run_widget, self.environment_name)

        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.netcdf_handle = None
        self.acquiring = False
        self.plot_data_items = {}
        self._force_plot_time = np.array([])
        self._force_plot_measured = np.array([])
        self._force_plot_target = np.array([])

        self.complete_ui()
        self.connect_callbacks()
        self._restore_last_used_settings()

    def complete_ui(self):
        self.definition_widget.sweep_type_selector.addItems(['Linear', 'Logarithmic'])
        self.definition_widget.direction_selector.addItems(['Up', 'Down'])
        self.definition_widget.sweep_type_selector.currentTextChanged.connect(
            self.update_sweep_rate_units)
        self.update_sweep_rate_units(self.definition_widget.sweep_type_selector.currentText())
        self.definition_widget.adaptive_tracking_bandwidth_checkbox.toggled.connect(
            self.update_tracking_bandwidth_enabled)
        self.update_tracking_bandwidth_enabled(
            self.definition_widget.adaptive_tracking_bandwidth_checkbox.isChecked())
        self.definition_widget.feedforward_enabled_checkbox.toggled.connect(
            self.update_feedforward_enabled)
        self.update_feedforward_enabled(
            self.definition_widget.feedforward_enabled_checkbox.isChecked())
        plot_item = self.run_widget.force_plot.getPlotItem()
        plot_item.showGrid(True, True, 0.25)
        plot_item.enableAutoRange()
        plot_item.getViewBox().enableAutoRange(enable=True)
        self.plot_data_items['force'] = multiline_plotter(
            np.arange(2), np.zeros((2, 2)), widget=self.run_widget.force_plot,
            other_pen_options={'width': 1}, names=['Measured Force', 'Target Force'])

    def connect_callbacks(self):
        self.run_widget.select_file_button.clicked.connect(self.select_file)
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.definition_widget.feedforward_file_button.clicked.connect(self.select_feedforward_file)

    def update_sweep_rate_units(self, sweep_type_text: str):
        if sweep_type_text == 'Linear':
            self.definition_widget.sweep_rate_selector.setSuffix(' Hz/s')
        else:
            self.definition_widget.sweep_rate_selector.setSuffix(' oct/min')

    def update_tracking_bandwidth_enabled(self, adaptive_checked: bool):
        self.definition_widget.tracking_bandwidth_selector.setEnabled(not adaptive_checked)
        self.definition_widget.tracking_cycles_selector.setEnabled(adaptive_checked)

    def update_feedforward_enabled(self, enabled: bool):
        for widget_name in ('feedforward_learning_rate_selector', 'feedforward_bins_per_decade_selector',
                             'feedforward_min_selector', 'feedforward_max_selector',
                             'feedforward_trim_gain_max_selector', 'feedforward_file_selector',
                             'feedforward_file_button', 'feedforward_load_on_start_checkbox',
                             'feedforward_save_on_finish_checkbox'):
            getattr(self.definition_widget, widget_name).setEnabled(enabled)

    def select_file(self):
        filename, file_filter = QtWidgets.QFileDialog.getSaveFileName(
            self.run_widget, 'Select NetCDF File to Save Sine Force Control Data',
            filter='NetCDF File (*.nc4)')
        if filename == '':
            return
        self.run_widget.data_file_selector.setText(filename)

    def select_feedforward_file(self):
        filename, file_filter = QtWidgets.QFileDialog.getSaveFileName(
            self.definition_widget, 'Select Feedforward Map File (JSON) to Load/Save',
            filter='JSON File (*.json)', options=QtWidgets.QFileDialog.DontConfirmOverwrite)
        if filename == '':
            return
        self.definition_widget.feedforward_file_selector.setText(filename)

    def initialize_data_acquisition(self, data_acquisition_parameters: DataAcquisitionParameters):
        self.log('Initializing Data Acquisition')
        self.data_acquisition_parameters = data_acquisition_parameters
        channels = data_acquisition_parameters.channel_list
        self.definition_widget.force_channel_selector.clear()
        for index, channel in enumerate(channels):
            if channel.feedback_device is None:
                name = '{:} {:}{:}'.format(
                    '' if channel.channel_type is None else channel.channel_type,
                    channel.node_number, channel.node_direction)
                self.definition_widget.force_channel_selector.addItem(name, index)
        self.definition_widget.sample_rate_display.setValue(data_acquisition_parameters.sample_rate)
        self.definition_widget.stop_frequency_selector.setMaximum(
            data_acquisition_parameters.sample_rate / 2)
        self.definition_widget.start_frequency_selector.setMaximum(
            data_acquisition_parameters.sample_rate / 2)
        # Re-applied here (in addition to __init__) specifically so the
        # force-channel selection -- only populated above, once the channel
        # table is known -- gets restored too; harmless no-op for every
        # other field, which was already restored in __init__.
        self._restore_last_used_settings()

    def collect_environment_definition_parameters(self) -> SineForceControlParameters:
        return SineForceControlParameters.from_ui(self)

    def initialize_environment(self) -> AbstractMetadata:
        self.log('Initializing Environment Parameters')
        data = self.collect_environment_definition_parameters()
        if data.force_channel_index is None:
            raise ValueError('No force control channel selected!')
        if data.f_start <= 0 or data.f_stop <= 0:
            raise ValueError('Start/Stop frequency must be positive!')
        if data.max_drive_v >= data.abort_drive_v:
            raise ValueError('Abort Drive Limit must be greater than Max Drive (control limit)!')
        if data.initial_drive_v <= 0:
            raise ValueError('Initial Drive must be a small positive value '
                              '(a pure multiplicative controller cannot move '
                              'away from a zero starting amplitude).')
        if data.feedforward_enabled:
            if data.feedforward_min_v >= data.feedforward_max_v:
                raise ValueError('Feedforward Max must be greater than Feedforward Min!')
            if not (data.feedforward_min_v <= data.initial_drive_v <= data.feedforward_max_v):
                raise ValueError('Initial Drive must lie within [Feedforward Min, Feedforward Max] -- '
                                  'it seeds the feedforward map fallback estimate.')
            if data.feedforward_trim_gain_max <= 1.0:
                raise ValueError('Max Trim Ratio must be greater than 1.0!')
        # Warn (but don't block) if the control loop updates faster than the
        # tracking filter can settle -- the controller would then react to a
        # not-yet-settled amplitude estimate, causing the drive/force
        # amplitude to oscillate around the target instead of converging.
        # 5.0 here must match ForceTrackingEstimator's default
        # valid_settle_time_constants (see force_tracking_estimator.py).
        # With adaptive tracking bandwidth, use the *worst case* (narrowest)
        # bandwidth reached anywhere in the configured sweep -- i.e. at
        # whichever of f_start/f_stop is lower -- since that is where the
        # filter is slowest to settle.
        if data.adaptive_tracking_bandwidth:
            worst_case_bandwidth_hz = ForceTrackingEstimator.bandwidth_for_tracking_cycles(
                min(data.f_start, data.f_stop), data.tracking_cycles)
        else:
            worst_case_bandwidth_hz = data.tracking_bandwidth_hz
        settle_time_s = 5.0 / (2 * np.pi * worst_case_bandwidth_hz)
        if data.control_update_period_s < settle_time_s:
            QtWidgets.QMessageBox.warning(
                self.definition_widget, 'Control Update Period May Be Too Short',
                'Control Update Period ({:.4f} s) is shorter than the tracking '
                'filter\'s settle time (~{:.4f} s at {:.2f} Hz bandwidth{:}).\n\n'
                'The controller may react to a not-yet-settled amplitude estimate, '
                'causing the drive/force amplitude to oscillate around the target '
                'instead of converging.\n\n'
                'Consider increasing Control Update Period to at least {:.4f} s '
                '(with some margin), and/or increasing Tracking Bandwidth / Tracking '
                'Cycles, and/or reducing Controller Alpha for a more damped response.\n\n'
                'You can still proceed with the current values.'.format(
                    data.control_update_period_s, settle_time_s, worst_case_bandwidth_hz,
                    ' at the lowest swept frequency, adaptive' if data.adaptive_tracking_bandwidth else '',
                    settle_time_s))
        self.environment_parameters = data
        self._save_last_used_settings(data)
        return data

    def _apply_definition_settings(self, get):
        """Populates every definition-tab field from ``get(name, default)``.

        Shared by :meth:`retrieve_metadata` (reads a saved test's netCDF
        group attributes) and :meth:`_restore_last_used_settings` (reads a
        small local JSON file of the last-used settings, see that method)
        so the two persistence paths -- "reopen a saved test" and "just
        remember what I last typed" -- cannot drift apart into two
        different field lists.
        """
        widget = self.definition_widget
        widget.sweep_type_selector.setCurrentText(
            SWEEP_TYPE_INTERNAL_TO_UI[get('sweep_type', 'linear')])
        widget.start_frequency_selector.setValue(get('f_start', 20.0))
        widget.stop_frequency_selector.setValue(get('f_stop', 500.0))
        widget.sweep_rate_selector.setValue(get('sweep_rate', 10.0))
        widget.direction_selector.setCurrentText(
            DIRECTION_INTERNAL_TO_UI[get('direction', 'up')])
        widget.repeat_sweep_checkbox.setChecked(bool(get('repeat', False)))
        widget.num_sweeps_selector.setValue(int(get('num_sweeps', 0)))
        widget.alternate_direction_checkbox.setChecked(bool(get('alternate_direction', False)))
        channel_index = get('force_channel_index', None)
        if channel_index is not None:
            index = widget.force_channel_selector.findData(int(channel_index))
            if index >= 0:
                widget.force_channel_selector.setCurrentIndex(index)
        widget.target_force_selector.setValue(get('target_force', 5.0))
        widget.force_floor_selector.setValue(get('force_floor', 0.1))
        widget.tracking_bandwidth_selector.setValue(get('tracking_bandwidth_hz', 5.0))
        widget.adaptive_tracking_bandwidth_checkbox.setChecked(
            bool(get('adaptive_tracking_bandwidth', False)))
        widget.tracking_cycles_selector.setValue(float(get('tracking_cycles', 3.0)))
        widget.controller_alpha_selector.setValue(get('controller_alpha', 0.5))
        widget.control_update_period_selector.setValue(get('control_update_period_s', 0.1))
        widget.initial_drive_selector.setValue(get('initial_drive_v', 0.05))
        widget.max_drive_selector.setValue(get('max_drive_v', 1.25))
        widget.abort_drive_selector.setValue(get('abort_drive_v', 1.4))
        widget.max_drive_step_selector.setValue(get('max_drive_step_v', 0.02))
        widget.ramp_time_selector.setValue(get('ramp_time_s', 1.0))
        widget.pre_dwell_time_selector.setValue(get('pre_dwell_time_s', 0.0))
        widget.feedforward_enabled_checkbox.setChecked(bool(get('feedforward_enabled', False)))
        widget.feedforward_learning_rate_selector.setValue(float(get('feedforward_learning_rate', 0.2)))
        widget.feedforward_min_selector.setValue(float(get('feedforward_min_v', 0.05)))
        widget.feedforward_max_selector.setValue(float(get('feedforward_max_v', 1.25)))
        widget.feedforward_trim_gain_max_selector.setValue(float(get('feedforward_trim_gain_max', 3.0)))
        widget.feedforward_bins_per_decade_selector.setValue(float(get('feedforward_bins_per_decade', 10.0)))
        widget.feedforward_file_selector.setText(str(get('feedforward_file', '')))
        widget.feedforward_load_on_start_checkbox.setChecked(bool(get('feedforward_load_on_start', False)))
        widget.feedforward_save_on_finish_checkbox.setChecked(bool(get('feedforward_save_on_finish', False)))

    def retrieve_metadata(self, netcdf_handle: nc4._netCDF4.Dataset):
        group = netcdf_handle.groups[self.environment_name]
        self._apply_definition_settings(lambda key, default: getattr(group, key, default))

    def _restore_last_used_settings(self):
        """Populates the definition tab from the last successfully-started
        test's settings, if any were saved (see :meth:`_save_last_used_settings`).

        Deliberately silent/best-effort: a missing, corrupt, or
        outdated-format file must never prevent the UI from opening --
        it just falls back to the .ui file's own hardcoded defaults, same
        as if this feature did not exist.
        """
        try:
            with open(LAST_USED_SETTINGS_PATH, 'r') as f:
                saved = json.load(f)
        except (OSError, ValueError):
            return
        try:
            self._apply_definition_settings(lambda key, default: saved.get(key, default))
        except (KeyError, ValueError, TypeError) as exc:
            self.log('Could not restore last-used settings: {:}'.format(exc))

    def _save_last_used_settings(self, data: SineForceControlParameters):
        """Saves ``data`` (the just-validated settings for a test about to
        run -- see :meth:`initialize_environment`) as the "last used"
        settings, so the next time this environment's definition tab is
        opened it starts pre-filled with these instead of the .ui file's
        hardcoded defaults. Deliberately hooked on successful validation
        (not on every keystroke or on a timer) so only a configuration that
        was actually used for a real test is ever remembered -- e.g. a
        forgotten Sweep Type change mid-edit, as long as it is never
        actually run, will not silently become tomorrow's default.
        """
        try:
            os.makedirs(os.path.dirname(LAST_USED_SETTINGS_PATH), exist_ok=True)
            with open(LAST_USED_SETTINGS_PATH, 'w') as f:
                json.dump(vars(data), f, indent=2)
        except OSError as exc:
            self.log('Could not save last-used settings: {:}'.format(exc))

    def create_netcdf_file(self, filename):
        self.netcdf_handle = nc4.Dataset(filename, 'w', format='NETCDF4', clobber=True)
        group_handle = self.netcdf_handle.createGroup(self.environment_name)
        self.environment_parameters.store_to_netcdf(group_handle)

    def start_control(self):
        filename = self.run_widget.data_file_selector.text()
        if filename == '':
            QtWidgets.QMessageBox.critical(
                self.run_widget, 'Invalid File',
                'Please select a file in which to store Sine Force Control diagnostics')
            return
        if self.run_widget.autoincrement_checkbox.isChecked():
            path, ext = os.path.splitext(filename)
            from glob import glob
            index = len(glob(path + '*' + ext))
            filename = '{:}_{:04d}{:}'.format(path, index, ext)
        self.create_netcdf_file(filename)
        self.acquiring = True
        self.run_widget.start_test_button.setEnabled(False)
        self.run_widget.stop_test_button.setEnabled(True)
        self.controller_communication_queue.put(
            self.log_name, (GlobalCommands.START_ENVIRONMENT, self.environment_name))
        self.environment_command_queue.put(
            self.log_name, (GlobalCommands.START_ENVIRONMENT, None))
        self.controller_communication_queue.put(
            self.log_name, (GlobalCommands.AT_TARGET_LEVEL, self.environment_name))

    def stop_control(self):
        self.environment_command_queue.put(self.log_name, (GlobalCommands.STOP_ENVIRONMENT, None))

    def update_gui(self, queue_data):
        message, data = queue_data
        if message == 'diagnostic_update':
            (t, frequency, target, measured, relative_error, drive_amplitude,
             state, saturated, valid, feedforward_value, feedback_correction_pct,
             feedforward_confidence, learning_applied, sweep_direction, sweep_number) = data
            self.run_widget.frequency_display.setValue(frequency)
            self.run_widget.drive_amplitude_display.setValue(drive_amplitude)
            self.run_widget.state_display.setText(state)
            self.run_widget.saturated_checkbox.setChecked(bool(saturated))
            self.run_widget.estimator_valid_checkbox.setChecked(bool(valid))
            self.run_widget.sweep_info_display.setText('#{:} ({:})'.format(sweep_number, sweep_direction))
            if np.isfinite(feedforward_value):
                self.run_widget.feedforward_value_display.setValue(feedforward_value)
            if np.isfinite(feedback_correction_pct):
                self.run_widget.feedback_correction_display.setValue(feedback_correction_pct)
            if valid:
                self.run_widget.measured_force_display.setValue(measured)
                self.run_widget.relative_error_display.setValue(relative_error * 100.0)
                self._force_plot_time = np.append(self._force_plot_time, t)[-MAX_PLOT_SAMPLES:]
                self._force_plot_measured = np.append(self._force_plot_measured, measured)[-MAX_PLOT_SAMPLES:]
                self._force_plot_target = np.append(self._force_plot_target, target)[-MAX_PLOT_SAMPLES:]
                self.plot_data_items['force'][0].setData(self._force_plot_time, self._force_plot_measured)
                self.plot_data_items['force'][1].setData(self._force_plot_time, self._force_plot_target)
            if self.acquiring and self.netcdf_handle is not None:
                group = self.netcdf_handle.groups[self.environment_name]
                i = group.dimensions['control_updates'].size
                group.variables['time'][i] = t
                group.variables['instantaneous_frequency'][i] = frequency
                group.variables['force_target'][i] = target
                group.variables['force_amplitude_measured'][i] = measured if valid else np.nan
                group.variables['relative_force_error'][i] = relative_error if valid else np.nan
                group.variables['drive_amplitude_command'][i] = drive_amplitude
                group.variables['controller_state'][i] = state
                group.variables['controller_saturated'][i] = 1 if saturated else 0
                group.variables['estimator_valid'][i] = 1 if valid else 0
                group.variables['feedforward_value'][i] = feedforward_value
                group.variables['feedback_correction_pct'][i] = feedback_correction_pct
                group.variables['feedforward_confidence'][i] = feedforward_confidence
                group.variables['feedforward_learning_applied'][i] = 1 if learning_applied else 0
                group.variables['sweep_direction'][i] = sweep_direction
                group.variables['sweep_number'][i] = sweep_number
        elif message == 'aborted':
            self.acquiring = False
            self.run_widget.state_display.setText('ABORTED: {:}'.format(data))
            self.run_widget.start_test_button.setEnabled(True)
            self.run_widget.stop_test_button.setEnabled(False)
            QtWidgets.QMessageBox.critical(
                self.run_widget, 'Sine Force Control Aborted', str(data))
        elif message == 'finished':
            self.acquiring = False
            self.run_widget.start_test_button.setEnabled(True)
            self.run_widget.stop_test_button.setEnabled(False)
            if self.netcdf_handle is not None:
                self.netcdf_handle.close()
                self.netcdf_handle = None

    @staticmethod
    def create_environment_template(environment_name: str, workbook: openpyxl.workbook.workbook.Workbook):
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1, 1, 'Control Type')
        worksheet.cell(1, 2, 'Sine Force Control')
        rows = [
            ('Sweep Type', 'Linear or Logarithmic'),
            ('Start Frequency', 'Hz'),
            ('Stop Frequency', 'Hz'),
            ('Sweep Rate', 'Hz/s (linear) or octaves/min (logarithmic)'),
            ('Direction', 'Up or Down'),
            ('Repeat', '0 or 1'),
            ('Number of Sweeps', 'Only used if Repeat=1. 0 = unlimited'),
            ('Alternate Direction', 'Only used if Repeat=1. 0 or 1 -- up/down/up/down... sweep'),
            ('Force Channel Index', '0-based index into the channel table'),
            ('Target Force', 'N peak'),
            ('Force Floor', 'N peak'),
            ('Tracking Bandwidth', 'Hz. Ignored if Adaptive Tracking Bandwidth=1'),
            ('Adaptive Tracking Bandwidth', '0 or 1 -- scale bandwidth with instantaneous sweep frequency'),
            ('Tracking Cycles', 'Only used if Adaptive Tracking Bandwidth=1. Filter time constant in drive cycles'),
            ('Controller Alpha', '0 < alpha <= 1'),
            ('Control Update Period', 's'),
            ('Initial Drive', 'V peak'),
            ('Max Drive', 'V peak (control limit)'),
            ('Abort Drive', 'V peak (safety limit)'),
            ('Max Drive Step', 'V peak per update'),
            ('Ramp Time', 's'),
            ('Pre-Sweep Dwell Time', 's'),
            ('Feedforward Enabled', '0 or 1 -- layer a learned frequency-dependent feedforward map on top of the fast loop'),
            ('Feedforward Learning Rate', '0 < rate <= 1, steady-state per-bin learning rate'),
            ('Feedforward Min', 'V peak, hard lower clamp on learned values'),
            ('Feedforward Max', 'V peak, hard upper clamp on learned values'),
            ('Feedforward Max Trim Ratio', '> 1.0, max multiplicative trim the fast loop may apply on top of the feedforward value'),
            ('Feedforward Bins Per Decade', 'log-frequency resolution of the learned curve'),
            ('Feedforward Map File', 'Path to JSON file for load/save, or empty for none'),
            ('Feedforward Load On Start', '0 or 1'),
            ('Feedforward Save On Finish', '0 or 1'),
        ]
        for i, (label, note) in enumerate(rows, start=2):
            worksheet.cell(i, 1, label)
            worksheet.cell(i, 2, '#')
            worksheet.cell(i, 4, 'Note: ' + note)

    def set_parameters_from_template(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        self.definition_widget.sweep_type_selector.setCurrentText(str(worksheet.cell(2, 2).value))
        self.definition_widget.start_frequency_selector.setValue(float(worksheet.cell(3, 2).value))
        self.definition_widget.stop_frequency_selector.setValue(float(worksheet.cell(4, 2).value))
        self.definition_widget.sweep_rate_selector.setValue(float(worksheet.cell(5, 2).value))
        self.definition_widget.direction_selector.setCurrentText(str(worksheet.cell(6, 2).value))
        self.definition_widget.repeat_sweep_checkbox.setChecked(bool(int(worksheet.cell(7, 2).value)))
        self.definition_widget.num_sweeps_selector.setValue(int(worksheet.cell(8, 2).value))
        self.definition_widget.alternate_direction_checkbox.setChecked(bool(int(worksheet.cell(9, 2).value)))
        index = self.definition_widget.force_channel_selector.findData(int(worksheet.cell(10, 2).value))
        if index >= 0:
            self.definition_widget.force_channel_selector.setCurrentIndex(index)
        self.definition_widget.target_force_selector.setValue(float(worksheet.cell(11, 2).value))
        self.definition_widget.force_floor_selector.setValue(float(worksheet.cell(12, 2).value))
        self.definition_widget.tracking_bandwidth_selector.setValue(float(worksheet.cell(13, 2).value))
        self.definition_widget.adaptive_tracking_bandwidth_checkbox.setChecked(bool(int(worksheet.cell(14, 2).value)))
        self.definition_widget.tracking_cycles_selector.setValue(float(worksheet.cell(15, 2).value))
        self.definition_widget.controller_alpha_selector.setValue(float(worksheet.cell(16, 2).value))
        self.definition_widget.control_update_period_selector.setValue(float(worksheet.cell(17, 2).value))
        self.definition_widget.initial_drive_selector.setValue(float(worksheet.cell(18, 2).value))
        self.definition_widget.max_drive_selector.setValue(float(worksheet.cell(19, 2).value))
        self.definition_widget.abort_drive_selector.setValue(float(worksheet.cell(20, 2).value))
        self.definition_widget.max_drive_step_selector.setValue(float(worksheet.cell(21, 2).value))
        self.definition_widget.ramp_time_selector.setValue(float(worksheet.cell(22, 2).value))
        self.definition_widget.pre_dwell_time_selector.setValue(float(worksheet.cell(23, 2).value))
        self.definition_widget.feedforward_enabled_checkbox.setChecked(bool(int(worksheet.cell(24, 2).value)))
        self.definition_widget.feedforward_learning_rate_selector.setValue(float(worksheet.cell(25, 2).value))
        self.definition_widget.feedforward_min_selector.setValue(float(worksheet.cell(26, 2).value))
        self.definition_widget.feedforward_max_selector.setValue(float(worksheet.cell(27, 2).value))
        self.definition_widget.feedforward_trim_gain_max_selector.setValue(float(worksheet.cell(28, 2).value))
        self.definition_widget.feedforward_bins_per_decade_selector.setValue(float(worksheet.cell(29, 2).value))
        self.definition_widget.feedforward_file_selector.setText(str(worksheet.cell(30, 2).value or ''))
        self.definition_widget.feedforward_load_on_start_checkbox.setChecked(bool(int(worksheet.cell(31, 2).value)))
        self.definition_widget.feedforward_save_on_finish_checkbox.setChecked(bool(int(worksheet.cell(32, 2).value)))


class SineForceControlEnvironment(AbstractEnvironment):
    """Environment implementing the closed-loop force-controlled sine sweep."""

    def __init__(self,
                 environment_name: str,
                 queue_container: SineForceControlQueues,
                 acquisition_active: mp.Value,
                 output_active: mp.Value):
        super().__init__(
            environment_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active,
            output_active)
        self.queue_container = queue_container
        self.command_map[GlobalCommands.START_ENVIRONMENT] = self.run_environment

        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.measurement_channels = None
        self.output_channels = None
        self.force_channel_local_index = None

        self.startup = True
        self.aborted = False
        self.state = 'INIT'  # INIT -> RAMPING -> ACTIVE -> STOPPING -> (finished)
        self.output_gate = 0.0
        self.output_gate_target = 0.0
        self.output_gate_change = 0.0
        self.samples_since_control_update = 0
        self.control_update_samples = None
        self.elapsed_time = 0.0

        self.output_sweep_generator = None
        self.input_phase_generator = None
        self.estimator = None
        self.controller = None
        self.target_spec = None
        self.last_frequency = None
        self.last_result = None
        self._last_status = ControllerStatus.OK

        # Feedforward learning layer (see components/feedforward_map.py).
        # feedforward_map stays None whenever feedforward is disabled --
        # every code path below that touches it is guarded on that, so a
        # disabled/misconfigured feedforward layer cannot affect the fast
        # loop at all.
        self.feedforward_map = None
        self.feedforward_map_published = None
        self._feedforward_committed_leg = 0
        self.total_drive_amplitude = 0.0
        self._last_feedforward_value = float('nan')
        self._last_feedback_correction_pct = float('nan')
        self._last_feedforward_confidence = float('nan')
        self._last_learning_applied = False
        self._last_sweep_direction = 'up'
        self._last_sweep_number = 1

    def initialize_data_acquisition_parameters(self, data_acquisition_parameters: DataAcquisitionParameters):
        self.log('Initializing Data Acquisition Parameters')
        self.data_acquisition_parameters = data_acquisition_parameters
        self.measurement_channels = [
            index for index, channel in enumerate(data_acquisition_parameters.channel_list)
            if channel.feedback_device is None]
        self.output_channels = [
            index for index, channel in enumerate(data_acquisition_parameters.channel_list)
            if not channel.feedback_device is None]

    def initialize_environment_test_parameters(self, environment_parameters: SineForceControlParameters):
        self.log('Initializing Environment Parameters')
        self.environment_parameters = environment_parameters
        self.force_channel_local_index = self.measurement_channels.index(
            environment_parameters.force_channel_index)

        self.output_sweep_generator = SineSweepGenerator(
            sample_rate=environment_parameters.output_sample_rate,
            sweep_type=environment_parameters.sweep_type,
            f_start=environment_parameters.f_start,
            f_stop=environment_parameters.f_stop,
            sweep_rate=environment_parameters.sweep_rate,
            direction=environment_parameters.direction,
            repeat=environment_parameters.repeat,
            num_sweeps=environment_parameters.num_sweeps,
            alternate_direction=environment_parameters.alternate_direction,
            pre_dwell_time=environment_parameters.pre_dwell_time_s)
        self.input_phase_generator = SineSweepGenerator(
            sample_rate=environment_parameters.sample_rate,
            sweep_type=environment_parameters.sweep_type,
            f_start=environment_parameters.f_start,
            f_stop=environment_parameters.f_stop,
            sweep_rate=environment_parameters.sweep_rate,
            direction=environment_parameters.direction,
            repeat=environment_parameters.repeat,
            num_sweeps=environment_parameters.num_sweeps,
            alternate_direction=environment_parameters.alternate_direction,
            pre_dwell_time=environment_parameters.pre_dwell_time_s)
        if environment_parameters.adaptive_tracking_bandwidth:
            initial_tracking_bandwidth_hz = ForceTrackingEstimator.bandwidth_for_tracking_cycles(
                environment_parameters.f_start, environment_parameters.tracking_cycles)
        else:
            initial_tracking_bandwidth_hz = environment_parameters.tracking_bandwidth_hz
        self.estimator = ForceTrackingEstimator(
            sample_rate=environment_parameters.sample_rate,
            tracking_bandwidth_hz=initial_tracking_bandwidth_hz)

        if environment_parameters.feedforward_enabled:
            # Feedforward mode: u_total = A_FF(f) * g. `self.controller` is
            # the *same*, completely unmodified ForceAmplitudeController
            # class, reused for the trim role -- reparametrized to regulate
            # a dimensionless gain g (centered on 1.0) instead of a voltage.
            # This is mathematically equivalent to the disabled-feedforward
            # case with the roles of "state" and "external multiplier" swapped
            # (see components/feedforward_map.py module docstring for why a
            # multiplicative composition is what's actually consistent with
            # this controller's own multiplicative/log-amplitude update law,
            # and why an additive trim starting at 0 cannot work with it).
            self.feedforward_map = FeedforwardMap(
                f_min=min(environment_parameters.f_start, environment_parameters.f_stop),
                f_max=max(environment_parameters.f_start, environment_parameters.f_stop),
                initial_estimate=environment_parameters.initial_drive_v,
                value_min=environment_parameters.feedforward_min_v,
                value_max=environment_parameters.feedforward_max_v,
                bins_per_decade=environment_parameters.feedforward_bins_per_decade,
                learning_rate=environment_parameters.feedforward_learning_rate,
                max_relative_step_per_update=FEEDFORWARD_MAX_RELATIVE_STEP_PER_UPDATE,
                outlier_reject_ratio=FEEDFORWARD_OUTLIER_REJECT_RATIO,
                outlier_reject_min_observations=FEEDFORWARD_OUTLIER_REJECT_MIN_OBSERVATIONS,
                outlier_reject_persistence=FEEDFORWARD_OUTLIER_REJECT_PERSISTENCE,
                max_observations_cap=FEEDFORWARD_MAX_OBSERVATIONS_CAP,
                separate_direction=FEEDFORWARD_SEPARATE_DIRECTION)
            if environment_parameters.feedforward_load_on_start and environment_parameters.feedforward_file:
                try:
                    self.feedforward_map.load(environment_parameters.feedforward_file)
                    self.log('Loaded feedforward map from {:}'.format(environment_parameters.feedforward_file))
                except (OSError, ValueError, KeyError) as exc:
                    self.log('Could not load feedforward map ({:}) -- starting from an empty map: {:}'.format(
                        environment_parameters.feedforward_file, exc))
            # self.feedforward_map accumulates learning every control update
            # (the "write" side). self.feedforward_map_published is a frozen
            # snapshot of it that composition actually reads from (the
            # "read" side), refreshed only at sweep-leg boundaries (see
            # _update_feedforward_and_compose). This split is required, not
            # just an optimization: composing directly from the same map
            # instance that is simultaneously being learned from creates a
            # second, fast feedback loop (map learns from the composed
            # command; the composed command immediately depends on what was
            # just learned) layered on top of the fast loop's own feedback
            # -- verified to cause materially worse tracking (larger,
            # sustained oscillation) than feedforward being disabled
            # entirely, even on the very first sweep leg with an initially
            # empty map. Per-leg publishing also directly matches the
            # intended design (learn *across* sweeps, not within one -- see
            # module docstring) rather than being merely a stability patch.
            self.feedforward_map_published = copy.deepcopy(self.feedforward_map)
            self._feedforward_committed_leg = 0
            self.controller = ForceAmplitudeController(
                alpha=environment_parameters.controller_alpha,
                force_floor=environment_parameters.force_floor,
                max_drive_amplitude=environment_parameters.feedforward_trim_gain_max,
                max_amplitude_step=FEEDFORWARD_TRIM_GAIN_MAX_STEP,
                initial_drive_amplitude=1.0)
        else:
            self.feedforward_map = None
            self.feedforward_map_published = None
            self.controller = ForceAmplitudeController(
                alpha=environment_parameters.controller_alpha,
                force_floor=environment_parameters.force_floor,
                max_drive_amplitude=environment_parameters.max_drive_v,
                max_amplitude_step=environment_parameters.max_drive_step_v,
                initial_drive_amplitude=environment_parameters.initial_drive_v)
        self.target_spec = ConstantForceTarget(environment_parameters.target_force, force_unit='peak')

        self.control_update_samples = max(1, round(
            environment_parameters.control_update_period_s * environment_parameters.sample_rate))
        self.samples_since_control_update = 0
        self.elapsed_time = 0.0
        self.output_gate = 0.0
        self.output_gate_target = 0.0
        self.output_gate_change = 0.0
        self.state = 'INIT'
        self.aborted = False
        self.last_frequency = environment_parameters.f_start
        self.last_result = None
        self._last_status = ControllerStatus.OK
        self.total_drive_amplitude = environment_parameters.initial_drive_v
        self._last_feedforward_value = float('nan')
        self._last_feedback_correction_pct = float('nan')
        self._last_feedforward_confidence = float('nan')
        self._last_learning_applied = False
        self._last_sweep_direction = environment_parameters.direction
        self._last_sweep_number = 1

    def _update_feedforward_and_compose(self, ctrl_result):
        """Composes ``self.total_drive_amplitude`` (the actual command sent
        to the shaker) from this control update's ``ctrl_result``, and --
        when feedforward is enabled -- learns from it.

        With feedforward disabled, ``ctrl_result.drive_amplitude`` already
        *is* the drive amplitude in volts (unchanged fast-loop behavior);
        this method's clip/slew step is then a no-op (the controller already
        enforces the identical limits internally -- see
        ``initialize_environment_test_parameters``).

        With feedforward enabled, ``ctrl_result.drive_amplitude`` is instead
        the dimensionless trim gain ``g``, and
        ``u_total = A_FF(f) * g`` (see components/feedforward_map.py). The
        sweep leg/direction is read from ``input_phase_generator`` (the same
        generator already used to demodulate the incoming force block at
        this same frequency -- see module docstring on input/output phase
        generator pairing) so learning is tagged consistently with the
        frequency it actually observed.
        """
        frequency = self.last_frequency
        leg, direction = self.input_phase_generator.leg_and_direction(self.elapsed_time)
        self._last_sweep_direction = direction
        self._last_sweep_number = leg + 1

        if self.feedforward_map is None:
            self.total_drive_amplitude = ctrl_result.drive_amplitude
            self._last_feedforward_value = float('nan')
            self._last_feedback_correction_pct = float('nan')
            self._last_feedforward_confidence = float('nan')
            self._last_learning_applied = False
        else:
            # Publish this run's learning-so-far into the read-side map only
            # at sweep-leg boundaries (never mid-leg) -- see
            # initialize_environment_test_parameters for why composing
            # directly from the same instance being learned from would
            # otherwise create a destructive fast feedback loop.
            if leg != self._feedforward_committed_leg:
                self.feedforward_map_published = copy.deepcopy(self.feedforward_map)
                self._feedforward_committed_leg = leg

            composition = compose_drive_amplitude(
                self.feedforward_map_published, frequency, ctrl_result.drive_amplitude,
                self.total_drive_amplitude, self.environment_parameters.max_drive_v,
                self.environment_parameters.max_drive_step_v, direction=direction)
            self.total_drive_amplitude = composition.total_drive_amplitude

            # Anti-windup: resync the trim controller's own state to what was
            # actually achieved (see compose_drive_amplitude docstring) --
            # otherwise it keeps marching ahead of physical reality every
            # time max_drive_v/max_drive_step_v clamp the composed signal,
            # and winds up pinned near its own feedforward_trim_gain_max
            # ceiling/floor regardless of the true remaining error.
            achieved_trim_gain = min(max(composition.achieved_trim_gain, 0.0),
                                      self.controller.max_drive_amplitude)
            self.controller.drive_amplitude = achieved_trim_gain

            # Learn from what was actually achieved (post slew/clip) -- that
            # is what the *next* force measurement will actually reflect --
            # rather than a possibly still-limited requested value. Only
            # when trustworthy (fast loop's own ratio-law request was not
            # held/saturated, tracking estimator settled -- see
            # ControllerStatus and ForceTrackingResult.valid). Written into
            # feedforward_map (the write side) -- takes effect for
            # composition only once published at the next leg boundary
            # above.
            trust = ctrl_result.status is ControllerStatus.OK
            learn_result = self.feedforward_map.update(
                frequency, observed_value=self.total_drive_amplitude, trust=trust, direction=direction)

            self._last_feedforward_value = composition.feedforward_value
            self._last_feedback_correction_pct = (achieved_trim_gain - 1.0) * 100.0
            self._last_feedforward_confidence = self.feedforward_map_published.confidence(frequency, direction=direction)
            self._last_learning_applied = learn_result.updated

    def _publish_diagnostics(self, frequency):
        valid = self.last_result.valid if self.last_result is not None else False
        measured = self.last_result.amplitude if (self.last_result is not None and valid) else float('nan')
        saturated = self._last_status is ControllerStatus.SATURATED
        target = self.target_spec.evaluate(frequency)
        relative_error = ((measured - target) / target) if valid else float('nan')
        state_str = '{:}/{:}'.format(self.state, self._last_status.value)
        self.queue_container.gui_update_queue.put((self.environment_name, (
            'diagnostic_update',
            (self.elapsed_time, frequency, target, measured, relative_error,
             self.total_drive_amplitude, state_str, saturated, valid,
             self._last_feedforward_value, self._last_feedback_correction_pct,
             self._last_feedforward_confidence, self._last_learning_applied,
             self._last_sweep_direction, self._last_sweep_number))))

    def _trigger_abort(self, reason: str):
        self.log('SAFETY ABORT: {:}'.format(reason))
        self.aborted = True
        self.state = 'ABORTED'
        self.output_gate = 0.0
        self.output_gate_target = 0.0
        self.output_gate_change = 0.0
        # Immediately mute this environment's output.
        try:
            silence = np.zeros((len(self.output_channels), self.data_acquisition_parameters.samples_per_write))
            self.queue_container.data_out_queue.put((silence, True))
        except Exception:
            pass
        # Request the existing global hardware-stop mechanism -- no
        # environment-specific abort path is invented here.
        self.controller_communication_queue.put(self.log_name, (GlobalCommands.STOP_HARDWARE, None))
        self.queue_container.gui_update_queue.put((self.environment_name, ('aborted', reason)))

    def run_environment(self, data):
        if self.aborted:
            return  # Latched -- no automatic resumption.
        try:
            self._run_environment_step()
        except Exception:
            self._trigger_abort('Unhandled exception in control loop:\n' + traceback.format_exc())

    def _run_environment_step(self):
        if self.startup:
            self.startup = False
            self.state = 'RAMPING'
            self.output_gate_target = 1.0
            ramp_samples = max(1, int(self.environment_parameters.ramp_time_s
                                       * self.environment_parameters.output_sample_rate))
            self.output_gate_change = 1.0 / ramp_samples

        # Consume the newest available acquisition block, if any.
        try:
            acquisition_data, last_acquisition = self.queue_container.data_in_queue.get_nowait()
        except mp.queues.Empty:
            acquisition_data = None
            last_acquisition = False

        if acquisition_data is not None:
            force_samples = np.asarray(acquisition_data[self.force_channel_local_index], dtype=float)
            n = force_samples.shape[-1]
            _, freq_in, phase_in = self.input_phase_generator.generate_block(n, drive_amplitude=1.0)
            if self.environment_parameters.adaptive_tracking_bandwidth:
                target_bandwidth_hz = ForceTrackingEstimator.bandwidth_for_tracking_cycles(
                    float(freq_in[0]), self.environment_parameters.tracking_cycles)
                self.estimator.set_tracking_bandwidth(
                    max(target_bandwidth_hz, MIN_ADAPTIVE_TRACKING_BANDWIDTH_HZ))
            self.last_result = self.estimator.process_block(force_samples, phase_in)
            self.last_frequency = float(freq_in[-1])
            self.elapsed_time += n / self.environment_parameters.sample_rate

            self.samples_since_control_update += n
            if (self.state == 'ACTIVE'
                    and self.samples_since_control_update >= self.control_update_samples):
                self.samples_since_control_update -= self.control_update_samples
                target = self.target_spec.evaluate(self.last_frequency)
                measured = self.last_result.amplitude if self.last_result.valid else None
                ctrl_result = self.controller.update(target, measured, self.last_result.valid)
                self._last_status = ctrl_result.status
                self._update_feedforward_and_compose(ctrl_result)
                if self.total_drive_amplitude > self.environment_parameters.abort_drive_v:
                    self._trigger_abort(
                        'Commanded drive amplitude {:.4f} V exceeded abort limit {:.4f} V'.format(
                            self.total_drive_amplitude, self.environment_parameters.abort_drive_v))
                    return
            self._publish_diagnostics(self.last_frequency)

        # Generate and send the next output block if the output task is ready for one.
        if self.queue_container.data_out_queue.empty():
            n_out = self.data_acquisition_parameters.samples_per_write
            samples, freq_out, phase_out = self.output_sweep_generator.generate_block(
                n_out, drive_amplitude=self.total_drive_amplitude)

            gate = self._advance_gate(n_out)
            output_block = samples * gate
            full_output = np.zeros((len(self.output_channels), n_out))
            # Single-output-channel drive: broadcast to all configured output
            # channels driving the shaker (matches the common single-shaker
            # setup this environment targets; a MIMO extension is future work).
            full_output[:, :] = output_block

            last_signal = (self.state == 'STOPPING' and self.output_gate <= 0.0
                            and self.output_gate_target == 0.0)
            self.queue_container.data_out_queue.put((copy.deepcopy(full_output), last_signal))

            if self.state == 'RAMPING' and self.output_gate >= 1.0:
                self.state = 'ACTIVE'
            if last_signal:
                while not last_acquisition:
                    self.log('Waiting for Last Acquisition')
                    try:
                        acquisition_data, last_acquisition = self.queue_container.data_in_queue.get(timeout=1.0)
                    except mp.queues.Empty:
                        break
                self.shutdown()
                return

        self.queue_container.environment_command_queue.put(
            self.environment_name, (GlobalCommands.START_ENVIRONMENT, None))

    def _advance_gate(self, n_samples: int) -> np.ndarray:
        """Advances the startup/shutdown ramp gate by n_samples and returns
        the per-sample gate value (0..1) to multiply onto the output."""
        if self.output_gate_change == 0.0:
            return np.full(n_samples, self.output_gate)
        gate = self.output_gate + (np.arange(n_samples) + 1) * self.output_gate_change
        reached = np.nonzero(
            np.abs(gate - self.output_gate_target) < abs(self.output_gate_change))[0]
        if len(reached) > 0:
            gate[reached[0] + 1:] = self.output_gate_target
            self.output_gate = self.output_gate_target
            self.output_gate_change = 0.0
        else:
            self.output_gate = gate[-1]
        return np.clip(gate, 0.0, 1.0)

    def stop_environment(self, data):
        """Ramps the drive amplitude gate down to zero and stops (normal, non-abort stop)."""
        if self.aborted:
            return
        self.state = 'STOPPING'
        self.output_gate_target = 0.0
        ramp_samples = max(1, int(self.environment_parameters.ramp_time_s
                                   * self.environment_parameters.output_sample_rate))
        self.output_gate_change = -1.0 / ramp_samples

    def shutdown(self):
        self.log('Shutting Down Sine Force Control')
        if (self.feedforward_map is not None
                and self.environment_parameters.feedforward_save_on_finish
                and self.environment_parameters.feedforward_file):
            try:
                self.feedforward_map.save(self.environment_parameters.feedforward_file)
                self.log('Saved feedforward map to {:}'.format(self.environment_parameters.feedforward_file))
            except OSError as exc:
                self.log('Could not save feedforward map: {:}'.format(exc))
        self.queue_container.environment_command_queue.flush(self.environment_name)
        self.queue_container.gui_update_queue.put((self.environment_name, ('finished', None)))
        self.startup = True


def sine_force_control_process(environment_name: str,
                               input_queue: VerboseMessageQueue,
                               gui_update_queue: Queue,
                               controller_communication_queue: VerboseMessageQueue,
                               log_file_queue: Queue,
                               data_in_queue: Queue,
                               data_out_queue: Queue,
                               acquisition_active: mp.Value,
                               output_active: mp.Value):
    """Sine Force Control environment process function called by multiprocessing.

    Creates a SineForceControlEnvironment object and runs it. Mirrors the
    structure of ``time_environment.time_process``.
    """
    queue_container = SineForceControlQueues(input_queue,
                                             gui_update_queue,
                                             controller_communication_queue,
                                             data_in_queue,
                                             data_out_queue,
                                             log_file_queue)
    process_class = SineForceControlEnvironment(
        environment_name,
        queue_container,
        acquisition_active,
        output_active)
    process_class.run()
